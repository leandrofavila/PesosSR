import datetime
from datetime import date, datetime, timedelta
from ConDB import DB
from openpyxl import load_workbook
import pandas as pd
import calendar


class PesoSemanal:
    def __init__(self, path_xlsx):
        self.path_xlsx = path_xlsx

    @staticmethod
    def val_insere():
        first_day, last_day = PesoMensal.last_monday()
        range_dias = '{} - {}'.format((first_day.strftime("%d/%m/%y")), (last_day.strftime("%d/%m/%y")))

        mass = DB(range_dias.split(' - ')[0], range_dias.split(' - ')[0])

        print(range_dias.split(' - ')[0], range_dias.split(' - ')[0])

        df = pd.DataFrame([mass.lib_pcp(),
                           mass.mp_consumida(),
                           mass.peso_car_fechado()],
                          columns=[str(range_dias)],
                          index=['peso_lib_pcp', 'mp_comsumida', 'peso_car_fechado'])

        new_list = [range_dias,
                    df[range_dias]['peso_lib_pcp'],
                    df[range_dias]['mp_comsumida'],
                    df[range_dias]['peso_car_fechado']]
        return new_list


    def find_empty(self):
        book = load_workbook(self.path_xlsx)
        sheet = book['Pesos Semana']
        vazia = ''
        for rowNum in range(1, sheet.max_row):
            if sheet.cell(row=rowNum, column=1).value is None:
                vazia = sheet.cell(row=rowNum, column=1)
                break

        vazia = str(vazia)
        vazia = vazia.removeprefix("<Cell 'Pesos Semana'.").removesuffix(">")
        next_row = vazia[1:]
        for row in sheet[vazia: 'D{}'.format(next_row)]:
            for idx, cell in enumerate(row):
                cell.value = self.val_insere()[idx]
        book.save(self.path_xlsx)


class PesoMensal(PesoSemanal):
    def __init__(self, path_xlsx):
        super().__init__(path_xlsx)


    def find_col_empty(self):
        book = load_workbook(self.path_xlsx)
        sheet = book['Pesos Mês']
        col = 2
        empty = ''
        for rowNum in range(col, sheet.max_row):
            if sheet.cell(row=rowNum, column=col).value is None:
                empty = sheet.cell(row=rowNum, column=col)
                break
        return book, sheet, empty


    @staticmethod
    def last_monday():
        now = datetime.now().weekday()
        less_days = now % 7
        if less_days == 0:
            less_days = 7
        print(less_days)
        last_monday = datetime.now() - timedelta(days=less_days)
        last_day = last_monday + timedelta(days=4)
        #print(last_monday, last_day)

        return last_monday, last_day



    @staticmethod
    def trata_data_mes():
        if 1 <= datetime.now().day <= 7:
            month_now = datetime.now().month
            month = (int(month_now) - 1) if month_now != 1 else 12
            init_day = datetime.now().date().strftime("20%y-" + str(month).zfill(2) + "-01")
            y, m = list(map(int, [date.today().strftime('%Y'), date.today().strftime('%m')]))
            last_day_month = (calendar.monthrange(y, month)[1])
            final_day = (str(datetime.now().date().strftime("20%y-" + str(month).zfill(2) + "-"))
                         + str(last_day_month if last_day_month != 0 else 1))
            return init_day, final_day

        else:
            dt = PesoMensal.last_monday()
            return dt



    def monta_df(self):
        #print(datetime.now() + timedelta(days=4))
        try:
            di, df = self.trata_data_mes()  #2024-03-01 2024-03-31
            print(di, df)
        except TypeError:
            raise

        mass_month = DB(di, df)
        lis_to_save = [mass_month.lib_pcp(),
                       mass_month.mp_consumida(),
                       mass_month.peso_car_fechado(),
                       mass_month.horas_tot() - mass_month.horas_elev(),
                       mass_month.horas_elev(),
                       mass_month.horas_tot()]
        print(lis_to_save)
        return lis_to_save


    def val_insere_month(self):
        book, sheet, empty = self.find_col_empty()
        vazia = str(empty).removeprefix("<Cell 'Pesos Mês'.").removesuffix(">")
        print(vazia)
        for row in sheet[vazia: 'G{}'.format(vazia[1:])]:
            for idx, cell in enumerate(row):
                cell.value = self.monta_df()[idx]
        book.save(self.path_xlsx)



path = r"C:\Users\pcp03\Desktop\pesos.xlsx"
sem = PesoSemanal(path)
sem.find_empty()
mes = PesoMensal(path)
mes.val_insere_month()
